# -*- coding: utf-8 -*-
"""
Created on Tue Jul 15 10:39:30 2025

@author: GAVITAB
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

eRDR_path = r"I:\Systeme\MoBI\ILG_Testing\eRDR_PROD\2509"
SCoD_path = r"I:\Systeme\MoBI\ILG_Testing\SCoD_PROD\2509"
folderpath = r"I:\Systeme\MoBI\ILG_Testing\PROD_Output\2509\Output"

def merge_ILG(path1, path2, save_path):
    """
    Merges specified sheet data from matching subfolders in path1 and path2.
    Creates a new Excel file with eight sheets: 'eRDR_Principal', 'SCoD_Principal', 
    'eRDR_Repricing', 'SCoD_Repricing', 'eRDR_Liqui', 'SCoD_Liqui', 'eRDR_Income', 'SCoD_Income'.
    """
    # Verify paths
    if not (os.path.exists(path1) and os.path.exists(path2) and os.path.exists(save_path)):
        print("One or more paths do not exist. Please check.")
        return
    
    # Get matching subfolders
    subfolders1 = set(os.listdir(path1))
    subfolders2 = set(os.listdir(path2))
    matching_folders = subfolders1.intersection(subfolders2)

    for folder in matching_folders:
        eRDR_folder = os.path.join(path1, folder)
        SCoD_folder = os.path.join(path2, folder)

        try:
            # Initialize a new workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Define the sheet names
            sheets = ['Principal', 'Repricing', 'Liqui', 'Income']
            
            # Define custom styles for number and date formatting
            number_style = NamedStyle(name="number_style", number_format="#,##0.00")  # Thousands separator, no decimals
            date_style = NamedStyle(name="date_style", number_format="DD.MM.YYYY")  # Date format dd.mm.yyyy

            # Process each sheet type
            for sheet in sheets:
                # Process eRDR folder for the current sheet
                eRDR_file = next((f for f in os.listdir(eRDR_folder) if f.endswith('.xlsx')), None)
                if eRDR_file:
                    eRDR_path = os.path.join(eRDR_folder, eRDR_file)
                    eRDR_data = pd.read_excel(eRDR_path, sheet_name=f'Input_Focus_{sheet}')
                    
                    # Write to 'eRDR_<sheet>' sheet
                    ws_eRDR = wb.create_sheet(title=f'eRDR_{sheet}')
                    for row in dataframe_to_rows(eRDR_data, index=False, header=True):
                        ws_eRDR.append(row)
                    
                    # Convert the header row (row 1) to text
                    for cell in ws_eRDR["1:1"]:  # Header row
                        cell.value = str(cell.value)  # Convert header value to text
                        
                    # Apply number format and date format
                    for row in ws_eRDR.iter_rows(min_row=2):  # Data rows
                        for cell in row:
                            if isinstance(cell.value, (int, float)):  # Format numbers
                                cell.style = number_style
                            elif isinstance(cell.value, pd.Timestamp):  # Format date
                                cell.style = date_style
                            
                            # Format 'AssetLiability' column as text
                            if cell.column_letter == 'AssetLiability':  # If column is AssetLiability
                                cell.number_format = '@'  # Text format
                                cell.value = str(cell.value)  # Convert to string

                # Process SCoD folder for the current sheet
                SCoD_file = next((f for f in os.listdir(SCoD_folder) if f.endswith('.xlsx')), None)
                if SCoD_file:
                    SCoD_path = os.path.join(SCoD_folder, SCoD_file)
                    SCoD_data = pd.read_excel(SCoD_path, sheet_name=f'Input_Focus_{sheet}')
                    
                    # Write to 'SCoD_<sheet>' sheet
                    ws_SCoD = wb.create_sheet(title=f'SCoD_{sheet}')
                    for row in dataframe_to_rows(SCoD_data, index=False, header=True):
                        ws_SCoD.append(row)
                    
                    # Convert the header row (row 1) to text
                    for cell in ws_SCoD["1:1"]:  # Header row
                        cell.value = str(cell.value)  # Convert header value to text
                    
                    # Apply number format and date format
                    for row in ws_SCoD.iter_rows(min_row=2):  # Data rows
                        for cell in row:
                            if isinstance(cell.value, (int, float)):  # Format numbers
                                cell.style = number_style
                            elif isinstance(cell.value, pd.Timestamp):  # Format date
                                cell.style = date_style
                            
                            # Format 'AssetLiability' column as text
                            if cell.column_letter == 'AssetLiability':  # If column is AssetLiability
                                cell.number_format = '@'  # Text format
                                cell.value = str(cell.value)  # Convert to string

                # Process eRDR_Income (apply date filter based on path2)
                if sheet == 'Income':  # Process only for 'Income' sheet
                    # Get the unique ReferenceDate values from path2
                    SCoD_income_file = next((f for f in os.listdir(SCoD_folder) if f.endswith('.xlsx')), None)
                    if SCoD_income_file:
                        SCoD_income_path = os.path.join(SCoD_folder, SCoD_income_file)
                        income_data_path2 = pd.read_excel(SCoD_income_path, sheet_name='Input_Focus_Income')
                        princ_data_path2 = pd.read_excel(SCoD_income_path, sheet_name='Input_Focus_Principal')
                        unique_dates = princ_data_path2["ReferenceDate"].drop_duplicates()
                        
                        income_data_path2['CustYield%'] = income_data_path2['CustYield_Country'] / income_data_path2['AvBalFunding_Country'] * 12
                        income_data_path2['FundingYield%'] = income_data_path2['FundingYield_Country'] / income_data_path2['AvBalFunding_Country'] * 12
                        income_data_path2['LiquiSpread%'] = income_data_path2['LiquiSpread_Country'] / income_data_path2['AvBalFunding_Country'] * 12
                        
                        # Apply the filter for 'ReferenceDate' in SCoD_Income
                        filtered_SCoD_income = income_data_path2[income_data_path2["ReferenceDate"].isin(unique_dates)]
                        
                        # Remove the existing 'SCoD_Income' sheet if it exists
                        if 'SCoD_Income' in wb.sheetnames:
                            del wb['SCoD_Income']
                        
                        ws_SCoD_income = wb.create_sheet(title='SCoD_Income')
                        for row in dataframe_to_rows(filtered_SCoD_income, index=False, header=True):
                            ws_SCoD_income.append(row)
                            
                        for row in ws_SCoD_income.iter_rows(min_row=2):
                            for cell in row:
                                if isinstance(cell.value, (int, float)):
                                    cell.style = number_style
                                elif isinstance(cell.value, pd.Timestamp):
                                    cell.style = date_style
    
                        # Filter the eRDR_Income data in path1
                        eRDR_income_file = next((f for f in os.listdir(eRDR_folder) if f.endswith('.xlsx')), None)
                        if eRDR_income_file:
                            eRDR_income_path = os.path.join(eRDR_folder, eRDR_income_file)
                            income_data_path1 = pd.read_excel(eRDR_income_path, sheet_name='Input_Focus_Income')
                            
                            income_data_path1['CustYield%'] = income_data_path1['CustYield_Country'] / income_data_path1['AvBalFunding_Country'] * 12
                            income_data_path1['FundingYield%'] = income_data_path1['FundingYield_Country'] / income_data_path1['AvBalFunding_Country'] * 12
                            income_data_path1['LiquiSpread%'] = income_data_path1['LiquiSpread_Country'] / income_data_path1['AvBalFunding_Country'] * 12                            

                            # Apply the filter for 'ReferenceDate' in eRDR_Income
                            filtered_eRDR_income = income_data_path1[income_data_path1["ReferenceDate"].isin(unique_dates)]

                            # Remove the existing 'eRDR_Income' sheet if it exists
                            if 'eRDR_Income' in wb.sheetnames:
                                del wb['eRDR_Income']

                            # Write the filtered data to 'eRDR_Income' sheet
                            ws_eRDR_income = wb.create_sheet(title='eRDR_Income')
                            for row in dataframe_to_rows(filtered_eRDR_income, index=False, header=True):
                                ws_eRDR_income.append(row)
                            
                            # Convert the header row (row 1) to text
                            for cell in ws_eRDR_income["1:1"]:  # Header row
                                cell.value = str(cell.value)  # Convert header value to text
                            
                            # Apply number format and date format
                            for row in ws_eRDR_income.iter_rows(min_row=2):  # Data rows
                                for cell in row:
                                    if isinstance(cell.value, (int, float)):  # Format numbers
                                        cell.style = number_style
                                    elif isinstance(cell.value, pd.Timestamp):  # Format date
                                        cell.style = date_style
                                    
                                    # Format 'AssetLiability' column as text
                                    if cell.column_letter == 'AssetLiability':  # If column is AssetLiability
                                        cell.number_format = '@'  # Text format
                                        cell.value = str(cell.value)  # Convert to string

            # Save the workbook
            save_file = os.path.join(save_path, f"{folder}_merged.xlsx")
            wb.save(save_file)
            print(f"Successfully merged for folder: {folder}")

        except Exception as e:
            print(f"Error processing folder {folder}: {e}")
            
def sort_ILG(folderpath):
    """
    Process all Excel files in the folder by applying filter and sorting logic to all the eRDR and SCoD sheets.
    """
    # Verify that the folder path exists
    if not os.path.exists(folderpath):
        print("The folder path does not exist. Please check.")
        return
    
    # Get all Excel files in the folder (excluding temporary files starting with "~$")
    excel_files = [f for f in os.listdir(folderpath) if f.endswith('.xlsx') and not f.startswith('~$')]
    
    if not excel_files:
        print("No valid Excel files found in the folder.")
        return
    
    for file in excel_files:
        filepath = os.path.join(folderpath, file)
        
        try:
            # Load the workbook
            wb_merged = load_workbook(filepath)
        except Exception as e:
            # Skip files that cannot be opened
            print(f"Error opening file {file}: {e}. Skipping.")
            continue
        
        # Process both sheets (eRDR and SCoD)
        sheets_to_process = ['eRDR_Principal', 'SCoD_Principal', 'eRDR_Repricing', 'SCoD_Repricing', 'eRDR_Liqui', 'SCoD_Liqui', 'eRDR_Income', 'SCoD_Income']
        for sheet_name in sheets_to_process:
            if sheet_name not in wb_merged.sheetnames:
                print(f"Sheet '{sheet_name}' not found in file '{file}'. Skipping.")
                continue  # Skip if the sheet does not exist
            
            ws = wb_merged[sheet_name]
            
            # Convert the sheet to a pandas DataFrame for easier manipulation
            data = ws.values
            columns = next(data)  # First row as column headers
            df = pd.DataFrame(data, columns=columns)
            
            # Apply sorting logic
            try:
                df = df.sort_values(
                    by=["ReferenceDate", "AssetLiability", "PlanningAccount", "RA_BalanceSheetEntity", "Currency"],
                    ascending=[False, True, True, True, True]
                )
            except KeyError as e:
                print(f"Skipping sheet '{sheet_name}' in file '{file}' due to missing column: {e}.")
                # Skip if required columns are missing
                continue
            
            # Write the DataFrame back to the sheet
            for i, row in enumerate(df.itertuples(index=False), start=2):
                for j, value in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=value)
        
        # Save the processed file
        wb_merged.save(filepath)
        wb_merged.close()
        print(f"File sorted and saved: {file}")

def compare_ILG(folderpath):
    """
    Compare pairs of sheets (e.g., eRDR_Principal with SCoD_Principal) in Excel files, calculate differences, and save results.
    """
    # Ensure the folder exists
    if not os.path.exists(folderpath):
        raise ValueError("Input path does not exist.")

    # Sheet pairs to compare and generate "Check" sheets
    sheet_pairs = [
        ("eRDR_Principal", "SCoD_Principal", "Check_Principal"),
        ("eRDR_Repricing", "SCoD_Repricing", "Check_Repricing"),
        ("eRDR_Liqui", "SCoD_Liqui", "Check_Liqui"),
        ("eRDR_Income", "SCoD_Income", "Check_Income"),
    ]
    ignored_columns = {"SnapshotName", "CreationDate", "RunNumber"}

    for file_name in os.listdir(folderpath):
        # Skip temporary/hidden files
        if file_name.startswith("~$") or not file_name.endswith(".xlsx"):
            continue

        file_path = os.path.join(folderpath, file_name)
        try:
            wb = load_workbook(file_path)

            # Initialize a flag for file comparison and reasons
            diff_found = False
            reasons = []  # Collect reasons for differences

            for eRDR_sheet, SCoD_sheet, check_sheet in sheet_pairs:
                if eRDR_sheet in wb.sheetnames and SCoD_sheet in wb.sheetnames:
                    ws_eRDR = wb[eRDR_sheet]
                    ws_SCoD = wb[SCoD_sheet]

                    # Create or clear the 'Check' sheet
                    if check_sheet in wb.sheetnames:
                        ws_check = wb[check_sheet]
                        wb.remove(ws_check)
                    ws_check = wb.create_sheet(check_sheet)

                    diff_sum = 0  # Track the sum of numeric differences
                    numeric_diff_columns = set()  # Track columns with numeric differences
                    text_diff_columns = set()  # Track columns with text differences

                    # Get headers from eRDR sheet to match columns
                    headers = [cell.value for cell in ws_eRDR[1]]

                    # Compare all rows
                    for i, eRDR_row in enumerate(ws_eRDR.iter_rows(min_row=2), start=2):  # Skip headers
                        SCoD_row = ws_SCoD[i] if i <= ws_SCoD.max_row else [None] * len(headers)
                        for j, eRDR_cell in enumerate(eRDR_row, start=1):
                            header = headers[j - 1]  # Map column to header
                            SCoD_cell = SCoD_row[j - 1] if j <= len(SCoD_row) else None

                            eRDR_value = eRDR_cell.value
                            SCoD_value = SCoD_cell.value if SCoD_cell else None

                            # Handle ignored columns by header name
                            if header in ignored_columns:
                                ws_check.cell(i, j).value = eRDR_value  # Copy directly
                                ws_check.cell(i, j).number_format = eRDR_cell.number_format  # Preserve format
                                continue

                            # Numeric comparisons
                            if isinstance(eRDR_value, (int, float)) or isinstance(SCoD_value, (int, float)):
                                eRDR_value = eRDR_value or 0
                                SCoD_value = SCoD_value or 0
                                try:
                                    diff = float(eRDR_value) - float(SCoD_value)
                                except ValueError:
                                    diff = None

                                ws_check.cell(i, j).value = diff
                                ws_check.cell(i, j).number_format = eRDR_cell.number_format  # Preserve format
                                if diff is not None and abs(diff) > 0.0001:
                                    diff_found = True
                                    diff_sum += diff
                                    numeric_diff_columns.add(header)
                                    ws_check.cell(i, j).fill = PatternFill(
                                        start_color="FFC8C8", end_color="FFC8C8", fill_type="solid"
                                    )
                            else:
                                # Text comparisons
                                if eRDR_value != SCoD_value:
                                    diff_found = True
                                    text_diff_columns.add(header)
                                    ws_check.cell(i, j).fill = PatternFill(
                                        start_color="FFFF99", end_color="FFFF99", fill_type="solid"
                                    )
                                ws_check.cell(i, j).value = eRDR_value or SCoD_value
                                ws_check.cell(i, j).number_format = eRDR_cell.number_format  # Preserve format

                    # Replace headers in Check sheet with eRDR headers
                    for col_idx, header in enumerate(headers, start=1):
                        ws_check.cell(1, col_idx).value = header

                    # Consolidate numeric differences message
                    if numeric_diff_columns:
                        reasons.append(
                            f"Numeric differences in sheet '{check_sheet}'"
                        )

                    # Consolidate text differences message
                    if text_diff_columns:
                        reasons.append(
                            f"Text differences in sheet '{check_sheet}', columns: {', '.join(text_diff_columns)}"
                        )

                    # Log if numeric differences are too high
                    if abs(diff_sum) >= 1:
                        diff_found = True
                        reasons.append(
                            f"Sum of numeric differences too high ({diff_sum}) in sheet '{check_sheet}'"
                        )

            # Determine the file status (_IDENTICAL or _NOT_IDENTICAL)
            new_name_suffix = "_IDENTICAL.xlsx" if not diff_found else "_NOT_IDENTICAL.xlsx"
            new_name = f"{file_name.rsplit('.', 1)[0]}{new_name_suffix}"

            # Print reasons for differences if NOT_IDENTICAL
            if diff_found:
                print(f"File: {file_name} marked as NOT_IDENTICAL due to the following reasons:")
                for reason in reasons:
                    print(f"  - {reason}")
            else:
                print(f"File: {file_name} marked as IDENTICAL.")

            # Rename the file
            new_path = os.path.join(folderpath, new_name)
            os.rename(file_path, new_path)

            # Save the updated workbook to the renamed file
            wb.save(new_path)
            print(f"Processed and saved: {new_name}")
            wb.close()
        except Exception as e:
            print(f"Error processing {file_name}: {e}")

merge_ILG(eRDR_path, SCoD_path, folderpath)
sort_ILG(folderpath)
compare_ILG(folderpath)
